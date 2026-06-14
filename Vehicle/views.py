from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.core.mail import send_mail
from django.conf import settings
from .models import Vehicle
from .forms import VehicleForm
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from datetime import date, timedelta
from django.contrib.auth import get_user_model
from django.core.paginator import Paginator

User = get_user_model()


def get_expiry_alerts():
    today = date.today()
    in_30_days = today + timedelta(days=30)
    alerts = []

    # بیمه
    for v in Vehicle.objects.filter(insurance_expiry__isnull=False, insurance_expiry__lte=in_30_days, insurance_expiry__gte=today):
        days_left = (v.insurance_expiry - today).days
        alerts.append({
            'vehicle': v,
            'type': 'Versicherung',
            'expiry': v.insurance_expiry,
            'days_left': days_left,
            'urgent': days_left <= 15,
        })

    # معاینه فنی
    for v in Vehicle.objects.filter(technical_expiry__isnull=False, technical_expiry__lte=in_30_days, technical_expiry__gte=today):
        days_left = (v.technical_expiry - today).days
        alerts.append({
            'vehicle': v,
            'type': 'HU/TÜV',
            'expiry': v.technical_expiry,
            'days_left': days_left,
            'urgent': days_left <= 15,
        })

    alerts.sort(key=lambda x: x['days_left'])
    return alerts


def send_expiry_emails():
    alerts = get_expiry_alerts()
    if not alerts:
        return

    manager_emails = list(
        User.objects.filter(groups__name='manager', email__isnull=False)
        .exclude(email='')
        .values_list('email', flat=True)
    )

    for alert in alerts:
        recipients = list(manager_emails)
        if alert['vehicle'].responsible and alert['vehicle'].responsible.email:
            if alert['vehicle'].responsible.email not in recipients:
                recipients.append(alert['vehicle'].responsible.email)
        if not recipients:
            continue

        subject = f"⚠️ Ablauf {alert['type']} - {alert['vehicle'].RegisterNumber}"
        message = (
            f"Fahrzeug: {alert['vehicle'].RegisterNumber} - {alert['vehicle'].get_Model_display()}\n"
            f"Typ: {alert['type']}\n"
            f"Ablaufdatum: {alert['expiry']}\n"
            f"Verbleibende Tage: {alert['days_left']}\n"
        )
        try:
            send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, recipients)
        except Exception as e:
            print(f"Email error: {e}")


@login_required
def vehicle_list(request):
    vehicles = Vehicle.objects.select_related('responsible').all()

    # SEARCH
    search = request.GET.get('search', '').strip()
    if search:
        vehicles = vehicles.filter(
            Q(RegisterNumber__icontains=search) |	
            Q(Description__icontains=search) |
            Q(responsible__username__icontains=search)
        )

    # FILTERS
    model_ = request.GET.get('model', '').strip()
    kind_ = request.GET.get('kind', '').strip()
    fuel_ = request.GET.get('fuel', '').strip()
    responsible_ = request.GET.get('responsible', '').strip()

    if model_:
        vehicles = vehicles.filter(Model=model_)
    if kind_:
        vehicles = vehicles.filter(Kind=kind_)
    if fuel_:
        vehicles = vehicles.filter(Fuel=fuel_)
    if responsible_:
        vehicles = vehicles.filter(responsible_id=responsible_)

    # SORTING
    sort = request.GET.get('sort', 'RegisterNumber')
    order = request.GET.get('order', 'asc')
    allowed_fields = ['RegisterNumber', 'Model', 'Kind', 'Fuel', 'mileage', 'insurance_expiry', 'technical_expiry']
    if sort in allowed_fields:
        vehicles = vehicles.order_by(f'-{sort}' if order == 'desc' else sort)

    query_params = request.GET.copy()
    query_params.pop('page', None)

    # PAGINATION
    paginator = Paginator(vehicles, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'vehicles': page_obj.object_list,
        'form': VehicleForm(),
        'alerts': get_expiry_alerts(),
        'users': User.objects.all(),
        'model_choices': Vehicle.ModelChoices.choices,
        'kind_choices': Vehicle.KindChoices.choices,
        'fuel_choices': Vehicle.FuelChoices.choices,
        'query_params': query_params.urlencode(),
        'current_sort': sort,
        'current_order': order,
        'page_obj': page_obj,
    }
    return render(request, 'Vehicle/vehicle_list.html', context)


@login_required
def vehicle_add(request):
    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('vehicle_list')
        vehicles = Vehicle.objects.all()
        return render(request, 'Vehicle/vehicle_list.html', {
            'form': form,
            'vehicles': vehicles,
            'show_modal': True,
            'alerts': get_expiry_alerts(),
            'users': User.objects.all(),
            'model_choices': Vehicle.ModelChoices.choices,
            'kind_choices': Vehicle.KindChoices.choices,
            'fuel_choices': Vehicle.FuelChoices.choices,
        })
    return redirect('vehicle_list')


@login_required
def vehicle_update(request, pk):
    vehicle = get_object_or_404(Vehicle, pk=pk)
    if request.method == 'POST':
        # تشخیص نوع request
        if request.content_type and 'multipart' in request.content_type:
            data = request.POST
        else:
            data = json.loads(request.body)

        # آپدیت فیلدها — فقط یک بار
        vehicle.Model = data.get('Model', vehicle.Model)
        vehicle.Kind = data.get('Kind', vehicle.Kind)
        vehicle.Fuel = data.get('Fuel', vehicle.Fuel)
        vehicle.mileage = data.get('mileage') or None
        vehicle.purchase_date = data.get('purchase_date') or None
        vehicle.insurance_expiry = data.get('insurance_expiry') or None
        vehicle.technical_expiry = data.get('technical_expiry') or None
        vehicle.Description = data.get('Description', vehicle.Description)
        responsible_id = data.get('responsible_id')
        vehicle.responsible_id = responsible_id if responsible_id else None

        # تصویر فقط اگه multipart بود
        if request.FILES and 'Image' in request.FILES:
            vehicle.Image = request.FILES['Image']

        vehicle.save()
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def vehicle_delete(request, pk):
    vehicle = get_object_or_404(Vehicle, pk=pk)
    if request.method == 'POST':
        vehicle.delete()
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'}, status=400)


@login_required
def vehicle_export_excel(request):
    vehicles = Vehicle.objects.select_related('responsible').all()

    search = request.GET.get('search', '').strip()
    model_ = request.GET.get('model', '').strip()
    kind_ = request.GET.get('kind', '').strip()
    fuel_ = request.GET.get('fuel', '').strip()

    if search:
        vehicles = vehicles.filter(Q(RegisterNumber__icontains=search) | Q(Description__icontains=search))
    if model_:
        vehicles = vehicles.filter(Model=model_)
    if kind_:
        vehicles = vehicles.filter(Kind=kind_)
    if fuel_:
        vehicles = vehicles.filter(Fuel=fuel_)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Vehicles'

    headers = ['Kennzeichen', 'Modell', 'Art', 'Kraftstoff', 'Km-Stand', 'Verantwortlicher', 'Versicherung', 'HU/TÜV']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='0d6efd')
        cell.alignment = Alignment(horizontal='center')

    for v in vehicles:
        ws.append([
            v.RegisterNumber,
            v.get_Model_display(),
            v.get_Kind_display() if v.Kind else '',
            v.get_Fuel_display() if v.Fuel else '',
            v.mileage or '',
            v.responsible.get_full_name() if v.responsible else '',
            str(v.insurance_expiry) if v.insurance_expiry else '',
            str(v.technical_expiry) if v.technical_expiry else '',
        ])

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=vehicles.xlsx'
    wb.save(response)
    return response


@login_required
def vehicle_export_pdf(request):
    vehicles = Vehicle.objects.select_related('responsible').all()

    search = request.GET.get('search', '').strip()
    model_ = request.GET.get('model', '').strip()
    kind_ = request.GET.get('kind', '').strip()

    if search:
        vehicles = vehicles.filter(Q(RegisterNumber__icontains=search) | Q(Description__icontains=search))
    if model_:
        vehicles = vehicles.filter(Model=model_)
    if kind_:
        vehicles = vehicles.filter(Kind=kind_)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=vehicles.pdf'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    elements = []

    data = [['Kennzeichen', 'Modell', 'Art', 'Kraftstoff', 'Km-Stand', 'Verantwortlicher', 'Versicherung', 'HU/TÜV']]
    for v in vehicles:
        data.append([
            v.RegisterNumber,
            v.get_Model_display(),
            v.get_Kind_display() if v.Kind else '',
            v.get_Fuel_display() if v.Fuel else '',
            str(v.mileage) if v.mileage else '',
            v.responsible.get_full_name() if v.responsible else '',
            str(v.insurance_expiry) if v.insurance_expiry else '',
            str(v.technical_expiry) if v.technical_expiry else '',
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 3),
    ]))

    elements.append(table)
    doc.build(elements)
    return response

