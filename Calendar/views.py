from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from .models import LeaveRequest
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from datetime import datetime, timedelta
from Settings.models import CompanySettings
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import openpyxl
import json

# Ensure CustomUser reference
CustomUser = get_user_model()


@login_required
def calendarView(request):
    return render(request, 'Calendar/calendar.html')

@login_required
def get_leaves(request):
    leaves = LeaveRequest.objects.filter(user=request.user)
    print("USER:", request.user)
    print("LEAVES COUNT:", leaves.count())

    events = []
    for leave in leaves:
        color = "#95A5A6"
        if leave.leave_type == "ill":
            if leave.status == "approved":
                color = "#E74C3C"
            elif leave.status == "pending":
                color = "#F1948A"
            else:
                color = "#7F8C8D"
        elif leave.leave_type == "holiday":
            if leave.status == "approved":
                color = "#3498DB"
            elif leave.status == "pending":
                color = "#AED6F1"
            else:
                color = "#7F8C8D"

        events.append({  #  اینجا باید با for هم‌تراز باشه
            "title": f"{leave.get_leave_type_display()} ({leave.get_status_display()})",
            "start": leave.start_date.strftime("%Y-%m-%d"),
            "end": (leave.end_date + timedelta(days=1)).strftime("%Y-%m-%d"),  #  یه روز اضافه
            "color": color,
            "extendedProps": {
                "leave_type": leave.leave_type,
                "status": leave.status,
                "id": leave.id,
            }
        })

    return JsonResponse(events, safe=False)  #  اینجا باید خارج از for باشه

@login_required
def add_leave(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        leave_type = data.get('leave_type')
        reason = data.get('reason', '')

        # چک تداخل با درخواست‌های قبلی
        overlapping = LeaveRequest.objects.filter(
            user=request.user,
            start_date__lte=end_date,
            end_date__gte=start_date
        ).exists()

        if overlapping:
            return JsonResponse({'status': 'error', 'message': 'You already have a leave request for this period!'})

        LeaveRequest.objects.create(
            user=request.user,
            start_date=start_date,
            end_date=end_date,
            leave_type=leave_type,
            reason=reason,
            status='pending'
        )
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def delete_leave(request, pk):
    leave = get_object_or_404(LeaveRequest, pk=pk, user=request.user)
    if request.method == 'POST':
        if leave.status == 'pending':
            leave.delete()
            return JsonResponse({'status': 'ok'})
        return JsonResponse({'status': 'error', 'message': 'Cannot delete approved/rejected leave'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def manager_leave_view(request):
    if not request.user.groups.filter(name='manager').exists():
        return render(request, 'Dashboard/no_access.html')

    pending = LeaveRequest.objects.filter(status='pending').order_by('start_date')
    processed = LeaveRequest.objects.exclude(status='pending').order_by('-start_date')

    context = {
        'pending': pending,
        'processed': processed,
    }
    return render(request, 'Calendar/manager_leave.html', context)

@login_required
def leave_report(request):
    if not request.user.groups.filter(name='manager').exists():
        return render(request, 'Dashboard/no_access.html')

    # فیلتر سال
    year = request.GET.get('year', datetime.now().year)
    
    # همه کارمندان
    users = CustomUser.objects.all()
    # حداکثر روز مرخصی از تنظیمات شرکت (یا مقدار پیش‌فرض 30)
    try:
        cs = CompanySettings.objects.first()
        max_days = getattr(cs, 'max_holiday_days', 30) or 30
    except:
        max_days = 30
    
    report_data = []
    for user in users:
        leaves = LeaveRequest.objects.filter(
            user=user,
            status='approved',
            start_date__year=year
        )
        
        holiday_days = sum(
            [(l.end_date - l.start_date).days + 1 for l in leaves.filter(leave_type='holiday')]
        )
        total_holiday_days = sum(item['holiday_days'] for item in report_data)
        ill_days = sum(
            [(l.end_date - l.start_date).days + 1 for l in leaves.filter(leave_type='ill')]
        )
        total_ill_days = sum(item['ill_days'] for item in report_data)
        pct = (holiday_days / max_days) * 100 if max_days else 0
        bar_color = '#dc3545' if pct > 83 else '#fd7e14' if pct > 50 else '#28a745'

        report_data.append({
            'user': user,
            'holiday_days': holiday_days,
            'ill_days': ill_days,
            'total_days': holiday_days + ill_days,
            'total_ill_days': total_ill_days,
            'total_holiday_days': total_holiday_days,
            'remaining': max_days - holiday_days,  # حداقل با مقدار تنظیمات شرکت
            'max_days': max_days,
            'bar_color': bar_color,


        })
        print("HOLIDAYDAYS:",total_holiday_days)
        print("ILLDAYS:",total_ill_days)

    # مرتب‌سازی بر اساس تعداد روزها
    report_data.sort(key=lambda x: x['total_days'], reverse=True)

    context = {
        'report_data': report_data,
        'year': int(year),
        'years': range(2023, datetime.now().year + 1),
        'total_ill_days': total_ill_days,
        'total_holiday_days': total_holiday_days,
        'max_days': max_days,

    }
    return render(request, 'Calendar/leave_report.html', context)


@login_required
def approve_leave(request, pk):
    if not request.user.groups.filter(name='manager').exists():
        return JsonResponse({'status': 'error'}, status=403)

    leave = get_object_or_404(LeaveRequest, pk=pk)
    if request.method == 'POST':
        data = json.loads(request.body)
        action = data.get('action')
        manager_note = data.get('manager_note', '')
        
        if action == 'approve':
            leave.status = 'approved'
            # آپدیت وضعیت Employee
            try:
                employee = leave.user.employee
                if leave.leave_type == 'holiday':
                    employee.status = 'vacation'
                elif leave.leave_type == 'ill':
                    employee.status = 'sick'
                employee.save()
            except:
                pass
                
        elif action == 'reject':
            leave.status = 'rejected'
            # برگردوندن به active
            try:
                employee = leave.user.employee
                employee.status = 'active'
                employee.save()
            except:
                pass

        leave.manager_note = manager_note
        leave.save()
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def leave_report_pdf(request):
    if not request.user.groups.filter(name='manager').exists():
        return JsonResponse({'status': 'error'}, status=403)

    year = request.GET.get('year', datetime.now().year)
    users = CustomUser.objects.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=leave_report_{year}.pdf'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20,
    )
    elements = []

    data = [['Employee', 'Holiday Days', 'Ill Days', 'Total Days', 'Remaining']]
    for user in users:
        leaves = LeaveRequest.objects.filter(
            user=user,
            status='approved',
            start_date__year=year
        )
        holiday_days = sum(
            [(l.end_date - l.start_date).days + 1 for l in leaves.filter(leave_type='holiday')]
        )
        ill_days = sum(
            [(l.end_date - l.start_date).days + 1 for l in leaves.filter(leave_type='ill')]
        )
        data.append([
            user.get_full_name() or user.username,
            holiday_days,
            ill_days,
            holiday_days + ill_days,
            30 - holiday_days,
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))

    elements.append(table)
    doc.build(elements)
    return response

@login_required
def leave_report_excel(request):
    if not request.user.groups.filter(name='manager').exists():
        return JsonResponse({'status': 'error'}, status=403)

    year = request.GET.get('year', datetime.now().year)
    users = CustomUser.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Leave Report'

    headers = ['Employee', 'Holiday Days', 'Ill Days', 'Total Days', 'Remaining']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='0d6efd')
        cell.alignment = Alignment(horizontal='center')

    for user in users:
        leaves = LeaveRequest.objects.filter(
            user=user,
            status='approved',
            start_date__year=year
        )
        holiday_days = sum(
            [(l.end_date - l.start_date).days + 1 for l in leaves.filter(leave_type='holiday')]
        )
        ill_days = sum(
            [(l.end_date - l.start_date).days + 1 for l in leaves.filter(leave_type='ill')]
        )
        ws.append([
            user.get_full_name() or user.username,
            holiday_days,
            ill_days,
            holiday_days + ill_days,
            30 - holiday_days,
        ])

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=leave_report_{year}.xlsx'
    wb.save(response)
    return response