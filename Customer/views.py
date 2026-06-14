from django.shortcuts import redirect, get_object_or_404, render
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table
from reportlab.platypus.tables import TableStyle
from .models import Customer
from .forms import CustomerForm
from Contract.models import Contract



@login_required
def customer_list(request):
    customers = Customer.objects.order_by()
    # SORTING
    sort = request.GET.get('sort', 'IDCustomer')
    order = request.GET.get('order', 'asc')

    allowed_fields = ['IDCustomer', 'Name', 'City', 'PostCode', 'MobileNumber1']

    if sort in allowed_fields:
        if order == 'desc':
            customers = customers.order_by(f'-{sort}')
        else:
            customers = customers.order_by(sort)

    # SEARCH
    search = request.GET.get('search')
    if search:
        customers = customers.filter(
            Q(IDCustomer__icontains=search) |
            Q(Name__icontains=search) |
            Q(City__icontains=search)
        )

    # FILTERS
    city = request.GET.get('city')
    postcode = request.GET.get('postcode')
    mobile = request.GET.get('mobile')

    if city:
        customers = customers.filter(City__icontains=city)

    if postcode:
        customers = customers.filter(PostCode__icontains=postcode)

    if mobile:
        customers = customers.filter(MobileNumber1__icontains=mobile)

    # PAGINATION
    paginator = Paginator(customers, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    query_params = request.GET.copy()
    query_params.pop('page', None)

    context = {
        'page_obj': page_obj,
        'customer_list': page_obj.object_list,
        'form': CustomerForm(),
        'query_params': '&'.join([f"{key}={value}" for key, value in request.GET.items() if key != 'page']),
        'current_sort': sort,
        'current_order': order,
    }
    return render(request, 'Customer/customer_list.html', context)

@login_required
def customer_add(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Customer added successfully!')  # ← اضافه شد
            print("MESSAGE ADDED")  # ← اضافه کن

            return redirect('customer_list')
        # فرم خطا داره
        customers = Customer.objects.all()
        paginator = Paginator(customers, 25)
        page_obj = paginator.get_page(1)
        return render(request, 'Customer/customer_list.html', {
            'form': form,
            'page_obj': page_obj,
            'customer_list': page_obj.object_list,
            'show_modal': True,
        })
    return redirect('customer_list')

@login_required
def customer_update(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        data = json.loads(request.body)
        customer.Name = data.get('Name', customer.Name)
        customer.City = data.get('City', customer.City)
        customer.PostCode = data.get('PostCode', customer.PostCode)
        customer.MobileNumber1 = data.get('MobileNumber1', customer.MobileNumber1)
        customer.save()
        messages.success(request, 'Customer updated successfully!')

        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer.delete()
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'}, status=400)


@login_required
def customer_contracts(request, pk):
    contracts = Contract.objects.filter(Customer_id=pk)
    data = []
    for c in contracts:
        data.append({
            'id': c.id,
            'IDContract': c.IDContract,
            'Name': c.Name,
            'Type': c.get_Type_display(),  # ← اضافه شد
            'DateRun': str(c.DateRun) if c.DateRun else '-',
            'Price': str(c.Price),
        })
    return JsonResponse({'contracts': data})

def customer_export_excel(request):
    customers = Customer.objects.all()

    # همون فیلترهای customer_list
    search = request.GET.get('search', '').strip()
    city = request.GET.get('city', '').strip()
    postcode = request.GET.get('postcode', '').strip()
    mobile = request.GET.get('mobile', '').strip()

    if search:
        customers = customers.filter(
            Q(IDCustomer__icontains=search) |
            Q(Name__icontains=search) |
            Q(City__icontains=search)
        )
    if city:
        customers = customers.filter(City__icontains=city)
    if postcode:
        customers = customers.filter(PostCode__icontains=postcode)
    if mobile:
        customers = customers.filter(MobileNumber1__icontains=mobile)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Customers'

    headers = ['ID Customer', 'Name', 'City', 'Post Code', 'Mobile']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='0d6efd')
        cell.alignment = Alignment(horizontal='center')

    for customer in customers:
        ws.append([
            customer.IDCustomer,
            customer.Name,
            customer.City,
            customer.PostCode,
            customer.MobileNumber1,
        ])

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=customers.xlsx'
    wb.save(response)
    return response

def customer_export_pdf(request):
    customers = Customer.objects.all()

    # همون فیلترهای customer_list
    search = request.GET.get('search', '').strip()
    city = request.GET.get('city', '').strip()
    postcode = request.GET.get('postcode', '').strip()
    mobile = request.GET.get('mobile', '').strip()

    if search:
        customers = customers.filter(
            Q(IDCustomer__icontains=search) |
            Q(Name__icontains=search) |
            Q(City__icontains=search)
        )
    if city:
        customers = customers.filter(City__icontains=city)
    if postcode:
        customers = customers.filter(PostCode__icontains=postcode)
    if mobile:
        customers = customers.filter(MobileNumber1__icontains=mobile)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=customers.pdf'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20,
    )
    elements = []

    data = [['ID Customer', 'Name', 'City', 'Post Code', 'Mobile']]
    for customer in customers:
        data.append([
            customer.IDCustomer,
            customer.Name,
            customer.City,
            customer.PostCode,
            customer.MobileNumber1,
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 3),
    ]))

    elements.append(table)
    doc.build(elements)
    return response