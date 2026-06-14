from decimal import Decimal

from django.shortcuts import redirect,get_object_or_404,render
from django.db.models import Q
from django.core.paginator import Paginator
from Customer.models import Customer
from .forms import ContractForm
from utils.views import FilterSortListView,ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from .models import Contract, ContractMaterial
from Warehouse.models import Material
from decimal import Decimal

@login_required
def contract_list(request):
    contracts = Contract.objects.order_by()

    # SEARCH
    search = request.GET.get('search')
    if search:
        contracts = contracts.filter(
            Q(IDContract__icontains=search) |
            Q(Name__icontains=search)
        )

    # FILTERS
    name = request.GET.get('name','').strip()
    customer = request.GET.get('customer','').strip()
    type_ = request.GET.get('type','').strip()
    date_run = request.GET.get('date_run','').strip()
    date_created = request.GET.get('date_created','').strip()

    if name:
        contracts = contracts.filter(Name__icontains=name)

    if customer:
        contracts = contracts.filter(Customer__Name__icontains=customer)
        print("AFTER CUSTOMER FILTER:", contracts.count())

    if type_:
        contracts = contracts.filter(Type=type_)

    if date_run:
        contracts = contracts.filter(DateRun=date_run)

    if date_created:
        contracts = contracts.filter(DateCreated=date_created)

    # SORTING
    sort = request.GET.get('sort', 'IDContract')
    order = request.GET.get('order', 'asc')

    allowed_fields = ['IDContract', 'Name', 'Customer', 'Type', 'Address', 'DateCreated', 'DateRun', 'Price', 'City']
    if sort in allowed_fields:
        contracts = contracts.order_by(f'-{sort}' if order == 'desc' else sort)
    
    # PAGINATION
    paginator = Paginator(contracts, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop('page', None)

    context = {
        'page_obj': page_obj,
        'contract_list': page_obj.object_list,
        'form': ContractForm(),
        'query_params': query_params.urlencode(),
        'current_sort': sort,
        'current_order': order,
        'customers': Customer.objects.all(),
        'all_materials': Material.objects.all(),

    }
    return render(request, 'Contract/contract_list.html', context)

@login_required
def contract_add(request):
    if request.method == 'POST':
        form = ContractForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Contract added successfully!')  # ← اضافه شد

            return redirect('contract_list')
        contracts = Contract.objects.all()
        paginator = Paginator(contracts, 25)
        page_obj = paginator.get_page(1)
        return render(request, 'Contract/contract_list.html', {
            'form': form,
            'page_obj': page_obj,
            'contract_list': page_obj.object_list,
            'show_modal': True,
            'customers': Customer.objects.all(),
        })
    return redirect('contract_list')

@login_required
def contract_update(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        data = json.loads(request.body)
        contract.Name = data.get('Name', contract.Name)
        contract.Customer_id = data.get('Customer_id', contract.Customer_id)
        contract.Type = data.get('Type', contract.Type)  # ← اضافه شد
        contract.Address = data.get('Address', contract.Address)
        contract.Price = data.get('Price', contract.Price)
        contract.City = data.get('City', contract.City)
        print("CONTRACT CITY:", contract.City)
        contract.save()
        messages.success(request, 'Contract updated successfully!')

        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def contract_delete(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        contract.delete()
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def contract_export_excel(request):
    contracts = Contract.objects.all()

    search = request.GET.get('search', '').strip()
    name = request.GET.get('name', '').strip()
    customer = request.GET.get('customer', '').strip()
    type_ = request.GET.get('type', '').strip()
    date_run = request.GET.get('date_run', '').strip()

    if search:
        contracts = contracts.filter(
            Q(IDContract__icontains=search) |
            Q(Name__icontains=search)
        )
    if name:
        contracts = contracts.filter(Name__icontains=name)
    if customer:
        contracts = contracts.filter(Customer__Name__icontains=customer)
    if type_:
        contracts = contracts.filter(Type=type_)
    if date_run:
        contracts = contracts.filter(DateRun=date_run)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contracts'

    # HEADER
    headers = ['ID Contract', 'Name', 'Customer', 'Type', 'Address','DateCreated', 'DateRun', 'Price']
    ws.append(headers)

    # استایل header
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='0d6efd')
        cell.alignment = Alignment(horizontal='center')

    # DATA
    for contract in contracts:
        ws.append([
            contract.IDContract,
            contract.Name,
            f"{contract.Customer.IDCustomer} - {contract.Customer.Name}" if contract.Customer else '',  # ← درست
            contract.get_Type_display(),
            contract.Address,
            contract.DateCreated,
            contract.DateRun,
            contract.Price
        ])

    # عرض ستون‌ها
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=contract.xlsx'
    wb.save(response)
    return response

@login_required
def contract_export_pdf(request):
    contracts = Contract.objects.all()

    # Apply same filters as the list view to the contracts queryset
    search = request.GET.get('search', '').strip()
    name = request.GET.get('name', '').strip()
    customer = request.GET.get('customer', '').strip()
    type_ = request.GET.get('type', '').strip()
    date_created = request.GET.get('date_created', '').strip()
    date_run = request.GET.get('date_run', '').strip()

    if search:
        contracts = contracts.filter(
            Q(IDContract__icontains=search) |
            Q(Name__icontains=search) |
            Q(Type__icontains=search) |
            Q(DateRun__icontains=search) |
            Q(DateCreated__icontains=search)
        )
    if name:
        contracts = contracts.filter(Name__icontains=name)
    if type_:
        contracts = contracts.filter(Type__icontains=type_)
    if date_run:
        contracts = contracts.filter(DateRun__icontains=date_run)
    if customer:
        # filter by related Customer ID or Name
        contracts = contracts.filter(
            Q(Customer__IDCustomer__icontains=customer) | Q(Customer__Name__icontains=customer)
        )
    if date_created:
        print("DATE CREATED SELECTED")
        contracts = contracts.filter(DateCreated__icontains=date_created)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=contracts.pdf'

    doc = SimpleDocTemplate(
        response, 
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20,
    )
    elements = []

    # DATA
    data = [['ID', 'Name', 'Customer', 'Type', 'Address', 'Date Created', 'Date Run', 'Price']]
    for contract in contracts:
        data.append([
            contract.IDContract,
            contract.Name[:20] if contract.Name else '',
            f"{contract.Customer.IDCustomer}" if contract.Customer else '',
            contract.get_Type_display(),
            contract.Address[:20] if contract.Address else '',
            str(contract.DateCreated) if contract.DateCreated else '',
            str(contract.DateRun) if contract.DateRun else '',
            f"€{contract.Price}",
        ])

    # TABLE
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),  # ← فونت کوچیک
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 3),  # ← padding کم
    ]))

    elements.append(table)
    doc.build(elements)
    return response


@login_required
def contract_materials(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    materials = ContractMaterial.objects.filter(contract=contract).select_related('material')
    
    return JsonResponse({
        'status': 'ok',
        'materials': [
            {
                'id': m.id,
                'material_id': m.material.id,
                'material_name': m.material.name,
                'material_type': m.material.get_type_display(),
                'required_kg': float(m.required_kg),
                'stock_kg': float(m.material.stock_kg),
                'stencil': m.stencil,
                'note': m.note,
            }
            for m in materials
        ]
    })


@login_required
def contract_material_add(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        data = json.loads(request.body)
        material = get_object_or_404(Material, pk=data.get('material_id'))
        required_kg = Decimal(str(data.get('required_kg', 0)))

        # چک موجودی
        warning = None
        if material.stock_kg < required_kg:
            warning = f'⚠️ Not enough stock! Available: {material.stock_kg} kg'

        # ساخت یا آپدیت ContractMaterial
        cm, created = ContractMaterial.objects.get_or_create(
            contract=contract,
            material=material,
            defaults={
                'required_kg': required_kg,
                'stencil': data.get('stencil', 'not_needed'),
                'note': data.get('note', ''),
            }
        )

        if not created:
            # برگشت مقدار قبلی به انبار
            material.stock_kg += cm.required_kg
            cm.required_kg = required_kg
            cm.stencil = data.get('stencil', cm.stencil)
            cm.note = data.get('note', cm.note)
            cm.save()

        # کسر از انبار
        material.stock_kg -= required_kg
        material.save()

        # ثبت تراکنش انبار
        from Warehouse.models import StockTransaction
        StockTransaction.objects.create(
            material=material,
            transaction_type='out',
            quantity_kg=required_kg,
            contract=contract,
            date=__import__('datetime').date.today(),
            note=f'Allocated to contract {contract.IDContract}',
            created_by=request.user
        )

        return JsonResponse({
            'status': 'ok',
            'warning': warning
        })
    return JsonResponse({'status': 'error'}, status=400)


@login_required
def contract_material_delete(request, pk):
    cm = get_object_or_404(ContractMaterial, pk=pk)
    if request.method == 'POST':
        # برگشت به انبار
        material = cm.material
        material.stock_kg += cm.required_kg
        material.save()

        # ثبت تراکنش برگشت
        from Warehouse.models import StockTransaction
        StockTransaction.objects.create(
            material=material,
            transaction_type='return',
            quantity_kg=cm.required_kg,
            contract=cm.contract,
            date=__import__('datetime').date.today(),
            note=f'Returned from contract {cm.contract.IDContract}',
            created_by=request.user
        )

        cm.delete()
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'}, status=400)